"""Genera Excel de inventario + manifest de integridad (paso 3)."""
from __future__ import annotations

import datetime as dt
import hashlib
import json
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from votaciones.config import (
    BASE, HEADERS_PDF, HEADERS_WARMUP, INDICES_DIR, PDFS_DIR, REPORTES_DIR,
    indices_dir_depto, reportes_dir_depto, safe_dir,
)
from votaciones.datos import (
    buscar_departamento, cargar_indices, transmisiones_del_depto,
)
from votaciones.descargar import construir_url


def _calcular_md5_sha256(p: Path) -> tuple[str, str]:
    md5 = hashlib.md5()
    sha = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            md5.update(chunk)
            sha.update(chunk)
    return md5.hexdigest(), sha.hexdigest()


def _validar_pdf(p: Path) -> tuple[str, int, bool]:
    if not p.exists():
        return "FALTA", 0, False
    try:
        size = p.stat().st_size
    except OSError:
        return "ERROR", 0, False
    if size == 0:
        return "VACIO", 0, False
    try:
        with p.open("rb") as f:
            magic = f.read(5)
    except OSError:
        return "ERROR", size, False
    if magic != b"%PDF-":
        return "NO_PDF", size, False
    if size < 1000:
        return "NO_PDF", size, True
    return "OK", size, True


def _ruta_local(t: dict, dept_name: str, mun_map: dict) -> Path:
    mun_name = mun_map.get(t["municipalityCode"], t["municipalityCode"])
    dest_dir = PDFS_DIR / safe_dir(dept_name) / safe_dir(f"{t['municipalityCode']}_{mun_name}")
    fname = f"z{t['idZoneCode']}_p{t['standCode']}_m{t['numberStand']}.pdf"
    return dest_dir / fname


def _migrar_legacy(dept_name: str) -> None:
    """Mueve archivos legacy planos a la subcarpeta por departamento.
    No borra nada, solo mueve si existe el archivo legacy y NO existe el destino."""
    import shutil
    nombre = safe_dir(dept_name)
    rep_depto = reportes_dir_depto(dept_name)
    legacy = [
        (REPORTES_DIR / f"inventario_{nombre}.xlsx", rep_depto / f"inventario_{nombre}.xlsx"),
        (REPORTES_DIR / f"manifest_{nombre}.jsonl",  rep_depto / f"manifest_{nombre}.jsonl"),
        (REPORTES_DIR / f"cambios_{nombre}.xlsx",    rep_depto / f"cambios_{nombre}.xlsx"),
    ]
    for src, dst in legacy:
        if src.exists() and not dst.exists():
            dst.parent.mkdir(parents=True, exist_ok=True)
            try:
                shutil.move(str(src), str(dst))
                print(f"  [migrado] {src.name} -> {dst}")
            except OSError as e:
                print(f"  [warning] No se pudo migrar {src.name}: {e}")


def _head_servidor(session: requests.Session, url: str, timeout: float = 30.0) -> dict:
    """Hace HEAD para capturar headers de integridad. Si HEAD da error o un 200
    sin ETag (bloqueo silencioso del CDN / Akamai sirviendo HTML default), usa
    GET con Range. En GET-Range el Content-Length viene de Content-Range."""
    try:
        r = session.head(url, headers=HEADERS_PDF, timeout=timeout, allow_redirects=True)
        usado_range = False
        # Caer al fallback si: error HTTP, O 200 sin ETag (Akamai bloqueando).
        # Un PDF legitimo de S3 SIEMPRE viene con ETag; si falta = no es el PDF.
        sin_etag = not (r.headers.get("ETag") or "").strip()
        if r.status_code not in (200, 301, 302) or sin_etag:
            r = session.get(url, headers={**HEADERS_PDF, "Range": "bytes=0-0"},
                            timeout=timeout, stream=True)
            r.close()
            usado_range = True
        h = r.headers
        # Para GET con Range, el Content-Length refleja solo el rango pedido
        # (ej. "1" para bytes=0-0). El tamaño real esta en Content-Range:
        # "bytes 0-0/56636" -> 56636. Lo extraemos para mantener consistencia.
        cl = h.get("Content-Length", "")
        if usado_range:
            cr = h.get("Content-Range", "")
            if "/" in cr:
                tamano_real = cr.rsplit("/", 1)[-1].strip()
                if tamano_real and tamano_real != "*":
                    cl = tamano_real
        etag = (h.get("ETag") or "").strip('"').lstrip("W/").strip('"')
        # Si ni con fallback obtuvimos ETag => CDN bloqueando
        http_status = r.status_code if etag else "BLOCKED_NO_ETAG"
        return {
            "http_status": http_status,
            "etag": etag,
            "last_modified": h.get("Last-Modified", ""),
            "version_id": h.get("x-amz-version-id", ""),
            "content_length": cl,
        }
    except requests.RequestException as e:
        return {"http_status": "ERROR", "etag": "", "last_modified": "",
                "version_id": "", "content_length": "", "error": str(e)[:80]}


def _respaldar_xlsx_existente(rep_depto: Path, dept_name: str) -> None:
    """Si ya existe inventario_<DEPTO>.xlsx o cambios_<DEPTO>.xlsx, copialo
    a reportes/<DEPTO>/backups/<nombre>_<timestamp>.xlsx antes de sobrescribir."""
    import shutil
    nombre = safe_dir(dept_name)
    backups = rep_depto / "backups"
    backups.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    for base in (f"inventario_{nombre}.xlsx", f"cambios_{nombre}.xlsx"):
        src = rep_depto / base
        if src.exists():
            dst = backups / f"{src.stem}_{stamp}.xlsx"
            shutil.copy2(src, dst)
            print(f"  [respaldo] {src.name} -> backups/{dst.name}")


def _ip_libre(session: requests.Session) -> tuple[bool, str]:
    """Verifica si la IP actual NO esta bloqueada por Akamai.
    Devuelve (libre, mensaje_diagnostico)."""
    session.cookies.clear()
    try:
        r = session.get(f"{BASE}/", headers=HEADERS_WARMUP, timeout=30)
    except requests.RequestException as e:
        return False, f"ERROR conexion: {e}"
    n_cookies = len(session.cookies)
    bloqueo = "Access Denied" in r.text
    if n_cookies > 0 and not bloqueo:
        return True, f"HTTP {r.status_code} cookies={n_cookies} bytes={len(r.content)} -> IP LIBRE"
    return False, f"HTTP {r.status_code} cookies={n_cookies} bytes={len(r.content)} -> BLOQUEADO"


def _esperar_cambio_de_red(session: requests.Session, intentos_max: int = 5) -> bool:
    """Pausa y le pide al usuario cambiar de IP (modo avion / hotspot / VPN).
    Verifica que la nueva IP no este bloqueada antes de continuar."""
    for intento in range(1, intentos_max + 1):
        print(f"\n{'!'*70}")
        print(f"  TU IP ESTA BLOQUEADA POR AKAMAI (intento {intento}/{intentos_max})")
        print(f"{'!'*70}")
        print(f"  Como cambiar IP en datos moviles (15 segundos):")
        print(f"    1. Activa MODO AVION en tu celular o PC con datos moviles")
        print(f"    2. Espera 5 segundos")
        print(f"    3. Desactiva modo avion")
        print(f"    4. Espera 5-10 segundos a que reconecte")
        print(f"  Cuando termines, vuelve aqui y presiona Enter.")
        try:
            input(f"\n  [Enter] cuando hayas cambiado de red  (Ctrl+C para cancelar): ")
        except (KeyboardInterrupt, EOFError):
            print(f"\n  [!] Cancelado por usuario.")
            return False
        print(f"  Verificando nueva IP...")
        libre, msg = _ip_libre(session)
        print(f"    {msg}")
        if libre:
            print(f"  [OK] IP nueva detectada, continuando reintento...")
            time.sleep(2)
            return True
        print(f"  [X] La nueva IP tambien esta bloqueada.")
        if intento < intentos_max:
            print(f"  Probemos otra vez: cambia a OTRA red (otro hotspot, VPN, etc.)")
    print(f"\n  [X] Despues de {intentos_max} intentos no logramos una IP libre.")
    return False


def _retry_bloqueados(session: requests.Session, filas_bloq: list[dict],
                       workers_original: int) -> None:
    """Reintenta los HEADs que dieron BLOCKED_NO_ETAG tras refrescar la sesion.
    Si la IP esta bloqueada, ofrece pausar para cambiar de red (modo avion)."""
    if not filas_bloq:
        return
    print(f"\n  [!] {len(filas_bloq)} archivos bloqueados por Akamai (sin ETag).")
    print(f"      Refrescando cookie ak_bmsc...")

    libre, msg = _ip_libre(session)
    print(f"      {msg}")
    if not libre:
        # IP marcada -> ofrecer cambio de red
        if not _esperar_cambio_de_red(session):
            print(f"      Saltando reintento; las filas bloqueadas quedan asi en el manifest.")
            return
    else:
        time.sleep(3)  # pausa para que Akamai baje el bot-score

    workers_retry = max(2, workers_original // 4)
    print(f"      Reintentando con {workers_retry} workers (vs {workers_original} originales)")

    t0 = time.time()
    pool = ThreadPoolExecutor(max_workers=workers_retry)
    try:
        futs = {pool.submit(_head_servidor, session, f["url"]): f for f in filas_bloq}
        done = 0
        total = len(futs)
        for fut in as_completed(futs):
            fila = futs[fut]
            srv_new = fut.result()
            # Solo sobrescribir si el reintento dio ETag (mejor que antes)
            if srv_new.get("etag"):
                fila["srv"] = srv_new
            done += 1
            if done % 50 == 0 or done == total:
                vel = done / (time.time() - t0) if time.time() > t0 else 0
                print(f"      retry {done}/{total}  ({vel:.1f}/s)")
    except KeyboardInterrupt:
        print(f"\n      [!] Cancelado, deteniendo retry...")
        pool.shutdown(wait=False, cancel_futures=True)
    finally:
        pool.shutdown(wait=False, cancel_futures=True)

    aun_bloq = sum(1 for f in filas_bloq if not f["srv"].get("etag"))
    recuperados = len(filas_bloq) - aun_bloq
    print(f"      Resultado: recuperados={recuperados}  aun_bloqueados={aun_bloq}")


def ejecutar(dept_code: str, *, workers: int = 10, sin_red: bool = False,
             out: Path | None = None, respaldar_xlsx: bool = False,
             prueba: int | None = None) -> int:
    tree, codes = cargar_indices()
    dept_name, mun_map = buscar_departamento(tree, dept_code)
    transmisiones = transmisiones_del_depto(codes, dept_code)

    print(f"\n>>> {dept_name} (codigo {dept_code})")
    print(f"    Mesas con PDF disponible: {len(transmisiones)}")
    if prueba is not None and prueba > 0:
        transmisiones = transmisiones[:prueba]
        print(f"    MODO PRUEBA: solo {len(transmisiones)} mesas")
    print(f"    Validando archivos en {PDFS_DIR / safe_dir(dept_name)}")
    if sin_red:
        print("    Consulta al servidor: NO (--sin-red)\n")
    else:
        print(f"    Consulta al servidor: SI (workers={workers})\n")

    if not transmisiones:
        print("No hay transmisiones para este departamento.")
        return 2

    # Carpetas por departamento (sin borrar lo existente, solo migrar)
    rep_depto = reportes_dir_depto(dept_name)
    rep_depto.mkdir(parents=True, exist_ok=True)
    _migrar_legacy(dept_name)
    if respaldar_xlsx:
        _respaldar_xlsx_existente(rep_depto, dept_name)

    wb = Workbook()
    ws = wb.active
    ws.title = safe_dir(dept_name)[:30] or "Inventario"
    columnas = [
        "departamento", "dept_code",
        "municipio_code", "municipio_nombre",
        "zona", "puesto", "mesa",
        "id_transmision", "id_stand",
        "expected_name",
        "url",
        "ruta_local",
        "descargado",
        "estado",
        "tamano_bytes",
        "magic_ok",
        "md5_archivo",
        "sha256_archivo",
        "servidor_etag",
        "etag_coincide",
        "servidor_last_modified",
        "servidor_version_id",
        "servidor_http_status",
        "servidor_content_length",
        "verificado_utc",
    ]
    ws.append(columnas)
    hf = Font(bold=True, color="FFFFFF")
    hp = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hp

    contadores = {"OK": 0, "FALTA": 0, "VACIO": 0, "NO_PDF": 0, "ERROR": 0}
    total_bytes = 0
    print("  [1/2] Validando archivos y calculando MD5+SHA-256 local...")
    filas: list[dict] = []
    for t in transmisiones:
        url = construir_url(t)
        path = _ruta_local(t, dept_name, mun_map)
        estado, size, magic_ok = _validar_pdf(path)
        contadores[estado] = contadores.get(estado, 0) + 1
        total_bytes += size
        md5_local = sha_local = ""
        if estado == "OK":
            try:
                md5_local, sha_local = _calcular_md5_sha256(path)
            except OSError:
                md5_local = sha_local = "ERROR"
        filas.append({
            "t": t, "url": url, "path": path, "estado": estado,
            "size": size, "magic_ok": magic_ok,
            "md5_local": md5_local, "sha_local": sha_local, "srv": {},
        })

    if not sin_red:
        print(f"  [2/2] Consultando {len(filas)} archivos al servidor...")
        session = requests.Session()
        try:
            session.get(f"{BASE}/", headers=HEADERS_WARMUP, timeout=60)
            n_cookies = len(session.cookies)
            print(f"        warmup OK  cookies={n_cookies}")
            if n_cookies == 0:
                # Akamai esta bloqueando esta IP. Ofrecer cambio de red.
                print(f"        [!] Akamai NO entrego cookie ak_bmsc -> tu IP esta BLOQUEADA.")
                if _esperar_cambio_de_red(session):
                    print(f"        [OK] Continuando con la nueva IP.")
                else:
                    print(f"        [!] Saltando consulta al servidor.")
                    sin_red = True
        except requests.RequestException as e:
            print(f"        WARMUP FAIL: {e}  -- columnas servidor_* quedaran vacias")
            sin_red = True

    if not sin_red:
        t0 = time.time()
        pool = ThreadPoolExecutor(max_workers=workers)
        try:
            futs = {pool.submit(_head_servidor, session, fila["url"]): fila for fila in filas}
            done = 0
            total = len(futs)
            for fut in as_completed(futs):
                fila = futs[fut]
                fila["srv"] = fut.result()
                done += 1
                if done % 100 == 0 or done == total:
                    dt_ = time.time() - t0
                    vel = done / dt_ if dt_ else 0
                    print(f"        {done}/{total}  ({vel:.1f}/s)")
        except KeyboardInterrupt:
            print(f"\n  [!] Cancelado por usuario. Deteniendo workers...")
            pool.shutdown(wait=False, cancel_futures=True)
            print(f"  [!] Se guardara lo procesado hasta ahora ({done}/{total} archivos).")
        finally:
            pool.shutdown(wait=False, cancel_futures=True)

        # Si Akamai bloqueo muchos, reintentar con re-warmup y menos workers
        filas_bloq = [f for f in filas if f["srv"].get("http_status") == "BLOCKED_NO_ETAG"]
        umbral_bloqueo = max(5, int(len(filas) * 0.05))  # >=5% o >=5 bloqueos
        if len(filas_bloq) >= umbral_bloqueo:
            _retry_bloqueados(session, filas_bloq, workers)

    verificado_utc = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    manifest_path = rep_depto / f"manifest_{safe_dir(dept_name)}.jsonl"
    with manifest_path.open("a", encoding="utf-8") as manifest:
        for fila in filas:
            t = fila["t"]
            srv = fila["srv"]
            etag_srv = srv.get("etag", "")
            etag_coincide: bool | str = ""
            if fila["md5_local"] and etag_srv and fila["md5_local"] != "ERROR":
                etag_coincide = (fila["md5_local"].lower() == etag_srv.lower())
            ws.append([
                dept_name,
                t["idDepartmentCode"],
                t["municipalityCode"],
                mun_map.get(t["municipalityCode"], ""),
                t["idZoneCode"],
                t["standCode"],
                t["numberStand"],
                t.get("idTransmissionCode", ""),
                t.get("idStand", ""),
                t["expectedName"],
                fila["url"],
                str(fila["path"]),
                fila["estado"] == "OK",
                fila["estado"],
                fila["size"],
                fila["magic_ok"],
                fila["md5_local"],
                fila["sha_local"],
                etag_srv,
                etag_coincide,
                srv.get("last_modified", ""),
                srv.get("version_id", ""),
                srv.get("http_status", ""),
                srv.get("content_length", ""),
                verificado_utc if srv else "",
            ])
            if srv and srv.get("etag"):
                manifest.write(json.dumps({
                    "verificado_utc": verificado_utc,
                    "url": fila["url"],
                    "md5_local": fila["md5_local"],
                    "sha256_local": fila["sha_local"],
                    "size_local": fila["size"],
                    "servidor": srv,
                }, ensure_ascii=False) + "\n")

    anchos = [14, 10, 14, 28, 8, 8, 8, 16, 14, 70, 110, 80, 12, 10, 14, 10,
              36, 66, 36, 12, 32, 36, 10, 12, 22]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill_falta = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    estado_col = columnas.index("estado") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[estado_col - 1].value != "OK":
            for cell in row:
                cell.fill = fill_falta

    ws2 = wb.create_sheet("resumen")
    ws2.append(["Departamento", dept_name])
    ws2.append(["Codigo", dept_code])
    ws2.append(["Mesas con PDF disponible", len(transmisiones)])
    ws2.append([])
    ws2.append(["Estado", "Cantidad", "Porcentaje"])
    total = len(transmisiones) or 1
    for estado in ("OK", "FALTA", "VACIO", "NO_PDF", "ERROR"):
        ws2.append([estado, contadores.get(estado, 0),
                    f"{contadores.get(estado, 0) / total * 100:.1f}%"])
    ws2.append([])
    ws2.append(["Total bytes descargados", total_bytes])
    ws2.append(["Total MB descargados", round(total_bytes / 1024 / 1024, 2)])
    etag_n = sum(1 for f in filas if f["srv"].get("etag"))
    ws2.append([])
    ws2.append(["Archivos con etag del servidor", etag_n])
    ws2.append(["Verificado UTC", verificado_utc])
    ws2.append(["Manifest", str(manifest_path)])
    for cell in ws2["A"]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 14

    if out is None:
        out = rep_depto / f"inventario_{safe_dir(dept_name)}.xlsx"
    wb.save(out)

    # Tambien guardar un extracto del catalogo del departamento en indices/<DEPTO>/
    ind_depto = indices_dir_depto(dept_name)
    ind_depto.mkdir(parents=True, exist_ok=True)
    extracto = {
        "departamento": dept_name,
        "codigo": dept_code,
        "generado_utc": verificado_utc,
        "total_mesas": len(transmisiones),
        "transmisiones": transmisiones,
    }
    (ind_depto / "transmisiones.json").write_text(
        json.dumps(extracto, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\nOK   : {contadores['OK']:>5}")
    print(f"FALTA: {contadores['FALTA']:>5}")
    print(f"VACIO: {contadores['VACIO']:>5}")
    print(f"NO_PDF:{contadores['NO_PDF']:>5}")
    print(f"ERROR: {contadores['ERROR']:>5}")
    print(f"\nTotal descargado: {total_bytes / 1024 / 1024:.1f} MB")
    print(f"Excel generado:   {out}")
    print(f"Manifest:         {manifest_path}")
    return 0


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python -m votaciones.inventario <CODE> [--workers K] [--sin-red] [--prueba N]")
        raise SystemExit(0)
    code = sys.argv[1].zfill(2)
    kw = {}
    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "--workers": kw["workers"] = int(sys.argv[i+1]); i += 2
        elif sys.argv[i] == "--sin-red": kw["sin_red"] = True; i += 1
        elif sys.argv[i] == "--prueba": kw["prueba"] = int(sys.argv[i+1]); i += 2
        else: i += 1
    raise SystemExit(ejecutar(code, **kw))
