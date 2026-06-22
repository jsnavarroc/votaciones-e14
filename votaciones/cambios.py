"""Detecta cambios en los E-14 entre corridas del paso 3 (paso 5)."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from votaciones.config import REPORTES_DIR, reportes_dir_depto, safe_dir
from votaciones.datos import buscar_departamento, cargar_indices


def _detectar_cambio(observaciones: list[dict]) -> dict:
    if len(observaciones) < 2:
        return {"cambio": False, "razones": [], "n": len(observaciones)}
    primera = observaciones[0]["servidor"]
    ultima = observaciones[-1]["servidor"]
    razones = []
    # Señales confiables de cambio real en el servidor (S3)
    for campo in ("etag", "version_id", "last_modified"):
        v1 = primera.get(campo, "")
        v2 = ultima.get(campo, "")
        if v1 and v2 and v1 != v2:
            razones.append(f"{campo}: {v1!r} -> {v2!r}")
    # content_length puede variar entre HEAD (tamaño completo) y GET-Range (1 byte)
    # del paso3. Solo lo consideramos si el ETag tambien cambio (cambio real).
    etag_cambio = any(r.startswith("etag:") for r in razones)
    cl1 = primera.get("content_length", "")
    cl2 = ultima.get("content_length", "")
    if etag_cambio and cl1 and cl2 and cl1 != cl2:
        razones.append(f"content_length: {cl1!r} -> {cl2!r}")
    return {
        "cambio": bool(razones),
        "razones": razones,
        "n": len(observaciones),
        "primera_utc": observaciones[0]["verificado_utc"],
        "ultima_utc": observaciones[-1]["verificado_utc"],
        "etag_primero": primera.get("etag", ""),
        "etag_ultimo": ultima.get("etag", ""),
        "version_primero": primera.get("version_id", ""),
        "version_ultimo": ultima.get("version_id", ""),
        "lm_primero": primera.get("last_modified", ""),
        "lm_ultimo": ultima.get("last_modified", ""),
    }


def ejecutar(dept_code: str, *, solo_cambios: bool = False) -> int:
    tree, _ = cargar_indices()
    dept_name, _ = buscar_departamento(tree, dept_code)
    rep_depto = reportes_dir_depto(dept_name)
    # Buscar manifest primero en la subcarpeta nueva, fallback al plano legacy
    manifest_path = rep_depto / f"manifest_{safe_dir(dept_name)}.jsonl"
    if not manifest_path.exists():
        legacy = REPORTES_DIR / f"manifest_{safe_dir(dept_name)}.jsonl"
        if legacy.exists():
            manifest_path = legacy
    if not manifest_path.exists():
        print(f"No existe {manifest_path}.")
        print(f"Necesitas correr 'Generar inventario' al menos DOS veces antes.")
        return 1

    obs_por_url: dict[str, list[dict]] = defaultdict(list)
    n_lineas = 0
    with manifest_path.open(encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea: continue
            try:
                r = json.loads(linea)
            except json.JSONDecodeError:
                continue
            obs_por_url[r["url"]].append(r)
            n_lineas += 1
    for lst in obs_por_url.values():
        lst.sort(key=lambda r: r["verificado_utc"])

    print(f"\n>>> {dept_name}")
    print(f"    Manifest:        {manifest_path}")
    print(f"    Lineas leidas:   {n_lineas}")
    print(f"    URLs unicas:     {len(obs_por_url)}")
    veces = [len(v) for v in obs_por_url.values()]
    if veces:
        print(f"    Observaciones por URL: min={min(veces)}  max={max(veces)}  avg={sum(veces)/len(veces):.1f}")
    if min(veces, default=0) < 2:
        urls_1 = sum(1 for v in obs_por_url.values() if len(v) < 2)
        print(f"    NOTA: {urls_1} URLs tienen solo 1 observacion -- no se pueden comparar")

    resultados = []
    for url, lst in obs_por_url.items():
        d = _detectar_cambio(lst)
        d["url"] = url
        resultados.append(d)
    cambios = [r for r in resultados if r["cambio"]]
    sin_cambios = [r for r in resultados if not r["cambio"] and r["n"] >= 2]
    solo_una = [r for r in resultados if r["n"] < 2]

    print(f"\n=== RESUMEN ===")
    print(f"  Con cambios detectados: {len(cambios)}")
    print(f"  Sin cambios (>=2 obs):  {len(sin_cambios)}")
    print(f"  Solo 1 observacion:     {len(solo_una)}")
    if cambios:
        print(f"\n=== ARCHIVOS QUE CAMBIARON (top 20) ===")
        for r in cambios[:20]:
            print(f"\n  {r['url']}")
            for razon in r["razones"]:
                print(f"    - {razon}")

    wb = Workbook()
    ws = wb.active
    ws.title = "cambios"
    cols = ["url", "n_obs", "primera_utc", "ultima_utc",
            "etag_primero", "etag_ultimo",
            "version_primero", "version_ultimo",
            "lm_primero", "lm_ultimo",
            "cambio", "razones"]
    ws.append(cols)
    hf = Font(bold=True, color="FFFFFF")
    hp = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for c in ws[1]:
        c.font = hf; c.fill = hp
    fill_cambio = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    lista = cambios if solo_cambios else (cambios + sin_cambios + solo_una)
    for r in lista:
        ws.append([
            r["url"], r["n"],
            r.get("primera_utc", ""), r.get("ultima_utc", ""),
            r.get("etag_primero", ""), r.get("etag_ultimo", ""),
            r.get("version_primero", ""), r.get("version_ultimo", ""),
            r.get("lm_primero", ""), r.get("lm_ultimo", ""),
            r["cambio"], " | ".join(r["razones"]),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[cols.index("cambio")].value is True:
            for cell in row:
                cell.fill = fill_cambio
    anchos = [110, 7, 22, 22, 36, 36, 36, 36, 32, 32, 9, 60]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    rep_depto.mkdir(parents=True, exist_ok=True)
    out = rep_depto / f"cambios_{safe_dir(dept_name)}.xlsx"
    wb.save(out)
    print(f"\nReporte: {out}")
    return 0 if not cambios else 2


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python -m votaciones.cambios <CODE> [--solo-cambios]")
        raise SystemExit(0)
    code = sys.argv[1].zfill(2)
    kw = {}
    if "--solo-cambios" in sys.argv:
        kw["solo_cambios"] = True
    raise SystemExit(ejecutar(code, **kw))
